#!/usr/bin/env python3
"""
validate_model.py — prototipo del motor de cálculo.

Hace DOS cosas:
  1) PRUEBA que la lógica reproduce el Excel: con los MISMOS inputs que la hoja
     "Vista B" (canceladas incluidas), el motor da NICA 6.064,82 · ALEX -373,56 ·
     MARE 310,15 · JACO 4.585,42 · TOTAL 10.586,82. Además el Ingreso Samavi por
     modelo coincide fila-a-fila con la columna derivada del Excel.
  2) Imprime los totales de PRODUCCIÓN (canceladas EXCLUIDAS, decisión jul 2026),
     que es lo que mostrarán las vistas SQL de 003_views.sql.

Uso: python3 scripts/validate_model.py ["STAG SAMAVI — Dashboard 2026.xlsx"]
"""
import sys
import datetime as dt
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "STAG SAMAVI — Dashboard 2026.xlsx"
YEAR, MONTHS = 2026, [1, 2, 3, 4]
CODES = ["1A_NICA", "4B_ALEX", "3G_MARE", "1A_JACO"]

# Referencia hoja Vista B (YTD ene-abr, CON canceladas)
EXPECTED = {
    "1A_NICA": dict(ingreso=17778.63, directo=13025.39, cuota=-6960.57, neto=6064.82),
    "4B_ALEX": dict(ingreso=13914.98, directo=5074.34,  cuota=-5447.90, neto=-373.56),
    "3G_MARE": dict(ingreso=11915.51, directo=4975.23,  cuota=-4665.08, neto=310.15),
    "1A_JACO": dict(ingreso=7952.48,  directo=7698.92,  cuota=-3113.50, neto=4585.42),
}
EXPECTED_TOTAL = dict(ingreso=51561.60, directo=30773.88, cuota=-20187.06, neto=10586.82)

wb = openpyxl.load_workbook(XLSX, data_only=True)
P = wb["⚙️ Parámetros"]
RAW = wb["_RAW_INGRESOS_2026"]


def c(r, col):
    return P.cell(r, col).value


# ── Bloque A: parámetros por propiedad ──
listings = {}
r = 11
while c(r, 1):
    fecha = c(r, 6)
    listings[c(r, 1)] = dict(
        modelo=(c(r, 5) or "").strip().lower().replace("comisión", "comision"),
        fecha_inicio=fecha.date() if isinstance(fecha, dt.datetime) else fecha,
        renta_base=c(r, 7) or 0, comision_pct=c(r, 8) or 0,
        limpieza=c(r, 11) or 0, suministros=c(r, 12) or 0, comunidad=c(r, 13) or 0,
        otros_fijos=sum((c(r, col) or 0) for col in (14, 15, 16, 17, 18, 19)),  # sin mobiliario_fin
    )
    r += 1

# ── Bloque B: SAMAVI_GEN recurrente €/mes ──
gen_mes = 0.0
r = 20
while c(r, 1) and not str(c(r, 1)).upper().startswith("TOTAL"):
    gen_mes += (c(r, 2) or 0); r += 1

# ── Bloque C: eventos ──
events = []
r = 50
while c(r, 1):
    events.append(dict(anio=int(c(r, 1)), mes=int(c(r, 2)), prop=c(r, 3), cat=c(r, 4), importe=c(r, 6) or 0))
    r += 1


def ev_sum(prop, cat):
    return sum(e["importe"] for e in events
              if e["prop"] == prop and e["cat"] == cat and e["anio"] == YEAR and e["mes"] in MONTHS)


def active_months(codigo):
    fi = listings[codigo]["fecha_inicio"]
    return [m for m in MONTHS if dt.date(YEAR, m, 1) >= dt.date(fi.year, fi.month, 1)]


def compute(include_canceled: bool):
    valid = {"confirmed", "checked_in", "checked_out"} | ({"canceled"} if include_canceled else set())
    ingreso = {k: 0.0 for k in listings}
    reservas = {k: 0 for k in listings}
    raw_col = {k: 0.0 for k in listings}   # columna "Ingreso P&L Samavi" del Excel (mismas filas)
    for row in RAW.iter_rows(min_row=2, values_only=True):
        codigo, status, checkin = row[4], row[6], row[1]
        if codigo not in listings or status not in valid or not isinstance(checkin, dt.datetime):
            continue
        if checkin.year != YEAR or checkin.month not in MONTHS:
            continue
        bruto, host_payout = row[10] or 0, row[13] or 0
        ing = bruto * listings[codigo]["comision_pct"] if listings[codigo]["modelo"] == "comision" else host_payout
        ingreso[codigo] += ing
        reservas[codigo] += 1
        raw_col[codigo] += (row[14] or 0)

    res = {}
    for k, p in listings.items():
        nm = len(active_months(k))
        gastos = (
            (-(p["renta_base"] * nm) + ev_sum(k, "RENTA"))
            - (p["limpieza"] * reservas[k])
            - (p["suministros"] * nm)
            - (p["comunidad"] * nm)
            - (p["otros_fijos"] * nm) + ev_sum(k, "OTROS")
        )
        res[k] = dict(ingreso=ingreso[k], directo=ingreso[k] + gastos, raw_col=raw_col[k], reservas=reservas[k])

    overhead = gen_mes * len(MONTHS) - sum(
        e["importe"] for e in events
        if e["prop"] == "SAMAVI_GEN" and e["anio"] == YEAR and e["mes"] in MONTHS)
    tot_ing = sum(ingreso.values())
    for k in res:
        res[k]["cuota"] = -overhead * (res[k]["ingreso"] / tot_ing) if tot_ing else 0
        res[k]["neto"] = res[k]["directo"] + res[k]["cuota"]
    return res


def near(a, b, tol=1.0):
    return abs(a - b) <= tol


# ── 1) PRUEBA: engine (con canceladas) == Excel Vista B ──
print("1) PRUEBA — el motor reproduce el Excel (mismos inputs, canceladas incluidas)")
print("=" * 66)
ref = compute(include_canceled=True)
ok = True
for k in CODES:
    row_ok = near(ref[k]["ingreso"], ref[k]["raw_col"])          # CASE por modelo == columna Excel
    vb_ok = all(near(ref[k][f], EXPECTED[k][f]) for f in ("ingreso", "directo", "cuota", "neto"))
    ok &= row_ok and vb_ok
    print(f"  {k}: CASE==Excel {'OK' if row_ok else 'DIFF'} · Vista B {'OK' if vb_ok else 'DIFF'} "
          f"(neto calc={ref[k]['neto']:.2f} exp={EXPECTED[k]['neto']:.2f})")
tot_neto = sum(ref[k]["neto"] for k in CODES)
ok &= near(tot_neto, EXPECTED_TOTAL["neto"], 2.0)
print(f"  TOTAL neto calc={tot_neto:.2f}  excel={EXPECTED_TOTAL['neto']:.2f}  {'OK' if ok else 'DIFF'}")
print("  " + ("✅ Motor validado contra Vista B\n" if ok else "❌ Revisar\n"))

# ── 2) PRODUCCIÓN: canceladas EXCLUIDAS (lo que darán las vistas SQL) ──
print("2) PRODUCCIÓN — canceladas EXCLUIDAS (003_views.sql)")
print("=" * 66)
prod = compute(include_canceled=False)
print(f"  {'Propiedad':10} {'Ingreso':>11} {'M.Directo':>11} {'Overhead':>11} {'M.Neto':>11}  reservas")
for k in CODES:
    p = prod[k]
    print(f"  {k:10} {p['ingreso']:11.2f} {p['directo']:11.2f} {p['cuota']:11.2f} {p['neto']:11.2f}  {p['reservas']}")
tt = {f: sum(prod[k][f] for k in CODES) for f in ("ingreso", "directo", "cuota", "neto")}
print(f"  {'TOTAL':10} {tt['ingreso']:11.2f} {tt['directo']:11.2f} {tt['cuota']:11.2f} {tt['neto']:11.2f}")
print(f"\n  Δ vs Excel (por canceladas excluidas): margen neto {tt['neto']-EXPECTED_TOTAL['neto']:+.2f} €")

sys.exit(0 if ok else 1)
